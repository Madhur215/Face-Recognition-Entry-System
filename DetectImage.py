import numpy as np
import pandas as pd
import os
import cv2
import openpyxl
from openpyxl import load_workbook
from datetime import datetime

names = { 0: 'n', 1 : 'm', 2: 'v', 3: 'l'}

class detect_image:

	def __init__(self, recordSheet, attendenceSheet):
		self.recordSheet = recordSheet
		self.attendenceSheet = attendenceSheet
		self.cascade = cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
		self.recognizer = cv2.face.LBPHFaceRecognizer_create()
		self.recognizer.read('trained.yml')
		self.students_marked = []
		self.recordWorkbook = load_workbook(filename=recordSheet)
		self.attendenceWorkbook = load_workbook(filename=attendenceSheet)

	def identify(self):
		cap = cv2.VideoCapture(0)
		f = 0
		pred_name = "None"
		while True:
			_, frame = cap.read()
			gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
			equalize = cv2.equalizeHist(gray_frame)
			image = cv2.medianBlur(equalize, 3)
			face = self.cascade.detectMultiScale(image, 1.1, 5)
			
			for x, y, w, h in face:
				cv2.rectangle(frame, (x, y), (x+w, y+h), (255, 255, 0), 4)
				label, conf = self.recognizer.predict(gray_frame[y:y+h, x:x+w])
				p_name = names.get(label)
				pred_name = p_name
				cv2.imshow("Recognizing Face", frame)
				flag = cv2.waitKey(1) & 0xFF
				f += 1
				if f > 15:
					cap.release()
					return p_name


	def writeEntry(self, p_name, branch, year):
		# num_students = len(self.students_marked)
		# print(self.students_marked)
		# print(branch, year)
		sheet = self.attendenceWorkbook.active
		sheet.insert_rows(idx=2, amount=1)
		sheet["A2"] = p_name
		sheet["B2"] = year
		sheet["C2"] = branch
		time = datetime.now()
		curr_time = time.strftime("%H:%M:%S")
		# print(curr_time)
		sheet["D2"] = curr_time
		self.attendenceWorkbook.save(self.attendenceSheet)

			
	def markEntry(self, p_name):
		if p_name in self.students_marked:
			print("Entry already marked!")
		else:
			self.students_marked.append(p_name)
			sheet = self.recordWorkbook.active
			for values in sheet.iter_rows(min_col = 1,
										max_col = 3,
										values_only = True):
				rec_name = values[0]
				if rec_name.lower() == p_name:
					branch = values[1]
					year = values[2]
					self.writeEntry(p_name, branch, year)


	def close_entry(self, p_name):
		if not p_name in self.students_marked:
			print("Entry not marked!")
		else:
			sheet = self.attendenceWorkbook.active
			index = 0
			for values in sheet.iter_rows(values_only = True):
				rec_name = values[0]
				index += 1
				if rec_name.lower() == p_name:
					time = datetime.now()
					curr_time = time.strftime("%H:%M:%S")
					sheet["E" + str(index)] = curr_time
					print("Entry closed!") 
					self.attendenceWorkbook.save(self.attendenceSheet)
					self.students_marked.remove(p_name)					
					break



